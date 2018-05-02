# Imports
import csv
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border, Side
import os
import xlsxwriter


def escape_txt(txt):
    # This function detects special characters and prefixes an apostrophe for to account for excel formating.
    special_chars = ['+', '=', '-', '/', '*', "'"]
    # Check to see if there is any value. If not, ignore.
    if txt:
        # Is first char in special char list.
        if txt[0] in special_chars:
            # print('Compensating (((((( {} ))))) by applying a prefix to accommodate excel formatting'.format(txt))
            logging.info('Compensating (((((( {} ))))) by applying a prefix to accommodate excel formatting'.format(txt))
            return "'" + txt
    return txt

def spread_sheet_creation(inputfile):
    # open csv file
    csv_file = open(inputfile)
    csv_ = csv.reader(csv_file)
    header_row = (next(csv_))
    #print(header_row)

    header_len = len(header_row) - 1 # figure out header lenght
    # create string for lenght of header
    position_header = xlsxwriter.utility.xl_col_to_name(header_len) + str(1)
    #csv_ = csv.reader(open('./Results.csv'))



    # Create an Excel workbook object
    wb = Workbook()
    # Create an Excel worksheet object
    ws = wb.active
    # Give the worksheet a title
    ws.title = "Results"

    # Go through each row of the header to write to output file
    row_idx = 1
    for cidx, val in enumerate(header_row):
        cell_idx = cidx + 1
        cell_name = '{}{}'.format(get_column_letter(cidx + 1), row_idx)
        cell = ws[cell_name]
        cell.value = escape_txt(val)
        cell.number_format = '@'


    # Go through each row in the csv in order to copy to XLSX with OPENPYXL library
    for ridx, row in enumerate(csv_):
        # Openpyxl starts row numbering at 1 so adjust the row index to match
        row_idx = ridx + 2
        # Go through each value in the csv row
        for cidx, val in enumerate(row):
            # Openpyxl starts column numbering at 1 so adjust the row index to match
            cell_idx = cidx + 1
            # Determine the excel cell name "A1", "B2", etc..
            # I have not seen this format used for a variable before.
            cell_name = '{}{}'.format(get_column_letter(cidx + 1), row_idx)
            # Create the cell object
            cell = ws[cell_name]
            # Set the value for the cell
            cell.value = escape_txt(val)
            # highight cell colors based True vs. False values
            if cell.value == "True" or cell.value == "PASS":
                cell.fill = PatternFill(fill_type='solid', fgColor='99ff66')
            elif cell.value == "False" or cell.value == "FAIL":
                cell.fill = PatternFill(fill_type='solid', fgColor='ff704d')
            # Set the cell format to text.  Don't ask my why @ means text but it does
            cell.number_format = '@'
            #print(cell)

    # Color and font adjustments to header
    # Define fill object
    HeaderFill = PatternFill(fill_type='solid', fgColor='ff0000')
    # define font objects for later application
    bold_font = Font(bold=True)
    white_font = Font(color='FFFFFF', italic=False, bold=True)

    # Loop through the cells in the first column and apply the bold formatting
    for cell in ws['A:A']:
        cell.font = bold_font
    # Loop through the cells in the first row and apply the white text formatting
    for cell in ws["1:1"]:
        cell.font = white_font
    # Loop through the cells in the first row and apply the border and red fill formatting
    try:
        for row in ws['A1':position_header]:
            for cell in row:
                cell.fill = HeaderFill
                cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                     right=Side(border_style='thin', color='FF000000'),
                                     bottom=Side(border_style='thin', color='FF000000'),
                                     left=Side(border_style='thin', color='FF000000'))
                # Freeze pane.  Column and row.
                c = ws['B2']
                # Freeze Panes
                ws.freeze_panes = c
    except:
        print("Please Close the output file")
    # print('Writing Cell: {}, Value: {}, Format: {}'.format(cell_name, cell.value, cell.number_format))
    # Save the file
    try:
        newfilename = inputfile.replace(".csv", ".xlsx")
        wb.save(filename=newfilename)
        #wb.save(filename='Result.xlsx')
    except:
        print("Please Close the output file")

    print("\nConversion of " + str(inputfile) + " to " + str(newfilename) + " completed.")

    csv_file.close()

if __name__ == '__main__':
    #spread_sheet_creation("numbers.csv")
    spread_sheet_creation()