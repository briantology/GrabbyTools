# Imports
import re
from collections import OrderedDict
import csv
from glob import glob
import logging
from os import listdir
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border, Side
import os
import time
import sys


def escape_txt(txt):
    # This function detects special characters and prefixes an apostrophe for to account for excel formating.
    special_chars = ['+', '=', '-', '/', '*', "'"]
    # Check to see if there is any value. If not, ignore.
    if txt:
        # Is first char in special char list.
        if txt[0] in special_chars:
            # print('Compensating (((((( {} ))))) by applying a prefix to accommodate excel formatting'.format(txt))
            logging.info('Compensating (((((( {} ))))) by applying a prefix to accommodate excel formatting'.format(txt))
            return "'" + txt
    return txt

def spread_sheet_creation():
    # try:
        # with open('Results.csv', 'w') as outputfile:
            # Create an object which operates like a regular writer but maps dictionaries onto Output rows
            # writer = csv.DictWriter(outputfile, fieldnames=headers, lineterminator='\n')  # define writer csv using the fieldnames columns
            # Figure out how many headers for later XLSX Fill style operation
            #column_count = len(headers)
            # Based on count, determine the column letter
            # maxcolumns = (get_column_letter(7)
            # Write Headers
            # writer.writeheader()
            # for i, j in devicesDictionary.items():
            #     writer.writerow(j)
    # except:
    #     print("Please Close the Netoutput.csv file and run the program again.")
    #     sys.exit()
    logging.info('############################################PROGRAM TERMINATED############################################')
    # open csv file
    csv_ = csv.reader(open('Results.csv', "w"))
    reader = csv.reader(csv_, delimiter=",")
    for row in reader:
        print(', '.join(row))



    # readheaders = next(reader)

    # Create an Excel workbook object
    wb = Workbook()
    # Create an Excel worksheet object
    ws = wb.active
    # Give the worksheet a title
    ws.title = "NetOutput"
    # Go through each row in the csv in order to copy to XLSX with OPENPYXL library
    for ridx, row in enumerate(csv_):
        # Openpyxl starts row numbering at 1 so adjust the row index to match
        row_idx = ridx + 1
        # Go through each value in the csv row
        for cidx, val in enumerate(row):
            # Openpyxl starts column numbering at 1 so adjust the row index to match
            cell_idx = cidx + 1
            # Determine the excel cell name "A1", "B2", etc..
            # I have not seen this format used for a variable before.
            cell_name = '{}{}'.format(get_column_letter(cidx + 1), row_idx)
            # Create the cell object
            cell = ws[cell_name]
            # Set the value for the cell
            cell.value = escape_txt(val)
            # Set the cell format to text.  Don't ask my why @ means text but it does
            cell.number_format = '@'
    # Define fill object
    HeaderFill = PatternFill(fill_type='solid', fgColor='ff0000')
    # define font objects for later application
    bold_font = Font(bold=True)
    white_font = Font(color='FFFFFF', italic=False, bold=True)
    # Loop through the cells in the first column and apply the bold formatting
    for cell in ws['A:A']:
        cell.font = bold_font
    # Loop through the cells in the first row and apply the white text formatting
    for cell in ws["1:1"]:
        cell.font = white_font
    # Loop through the cells in the first row and apply the border and red fill formatting
    try:
        for row in ws['A1':'A7']:
            for cell in row:
                cell.fill = HeaderFill
                cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                     right=Side(border_style='thin', color='FF000000'),
                                     bottom=Side(border_style='thin', color='FF000000'),
                                     left=Side(border_style='thin', color='FF000000'))
                # Freeze pane.  Column and row.
                c = ws['B2']
                # Freeze Panes
                ws.freeze_panes = c
    except:
        print("Please Close the NetOutput file")
    # print('Writing Cell: {}, Value: {}, Format: {}'.format(cell_name, cell.value, cell.number_format))
    # Save the file
    try:
        wb.save(filename='ResultOutput.xlsx')
    except:
        print("Please Close the NetOutput file")


spread_sheet_creation()