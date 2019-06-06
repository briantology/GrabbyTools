# Imports
# Changelog
# V1.1 - Updated code to offer fifth option to only collect show runs
#
#
#
#
#
#
# TODO Log
# Figure out why telnet fails on exception but works if DEVICE TYPE in input file is set to Cisco_ios_telnet
# Fix Failed output logs





import re
from collections import OrderedDict
import csv
from glob import glob
import logging
from os import listdir
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border, Side
import os
import time
import sys
import threading
import socket

#Todo Add ability to allow user to only download show run files or any of the specified files.


failed_connections_list = []

#Check for Log Directory if none detected, make one.

if os.path.isdir('Grabby_Logs'):
    pass
else:
    os.makedirs('Grabby_Logs')

# Set Logging Level
logging.basicConfig(filename=time.strftime("Grabby_Logs\%B_%d_%Y_%I.%M.%S_%p_Grabby_Log.txt"), level=logging.DEBUG, format='%(asctime)s:%(levelname)s:%(message)s')

logging.info('############################################PROGRAM AND FUNCTION INITIALIZATION############################################')


# Reg Ex Search Patterns for sh_run function
hostnameRE = re.compile(r'hostname\s(.+)')
snmpRE = re.compile(r'[sS]nmp-server community\s(.+)\s(.+)')
ipDomainRE = re.compile(r'ip domain name\s(.+)')
dnsRE = re.compile(r'ip name-server\s(.+)')
interfacetypeRE = re.compile(r'interface\s([SGFLVP].+)')
ipAddRE = re.compile(r'ip address\s(\d+\.\d+\.\d+\.\d+)\s(\d+\.\d+\.\d+\.\d+)')
intDescRE = re.compile(r'description\s(.+)')
ipChanGrpRE = re.compile(r'(channel-group) (\d+)(.+)')
ntpServerRE = re.compile(r'[nN][tT][pP] [sS]erver\s(.+)')
ipRouteRE = re.compile(r'ip route 0.0.0.0 0.0.0.0\s(.+)')
voicePortRE = re.compile(r'voice-port (.+)')
showRunRE = re.compile(r'(.+)(\s)(show run.txt)')
showIntRE1 = re.compile(r'(.+)(\s)(show interface.txt)')
showIntRE2 = re.compile(r'(.+)(\s)(show interface.txt)')
sccpRegex = re.compile(r'sccp ccm (\d+.\d+.\d+.\d+) identifier (\d)')

# Reg Ex Search Patterns for sh_int function
showMacRE = re.compile(r'.+(address is)\s(..............).+')

# Reg Ex Search Patterns for sh_ver function
showVerRE = re.compile(r'(Cisco)\s(.+)(Version\s)(\d.+),.+')
serialNumRE = re.compile(r'Processor board ID\s(.+)')
lastReloadReasonRE = re.compile(r'Last reload type:\s(.+)')
lastReloadTypeRE = re.compile(r'Last reload reason:\s(.+)')
lastReloadTimeRE = re.compile(r'System restarted at\s(.+)')
configRegisterRE = re.compile(r'Configuration register is\s(.+)')
modelTypeRE = re.compile(r'(Cisco)\s(.+)\s\(.+')
modelTypeRE2 = re.compile(r'(cisco ISR)(\d+)(.+)')
nvRamRE = re.compile(r'(\d+K)( bytes of non-volatile configuration memory.)')
nvRamRE2 = re.compile(r'(\d+K)( bytes of NVRAM.)')
flashRamRE = re.compile(r'(\d+K) (bytes of ATA System CompactFlash.+)')
flashRamRE2 = re.compile(r'(\d+K) (bytes of non-volatile configuration memory.+)')

# Reg Ex Search Patterns for file Inventory
fileshowVerRE2 = re.compile(r'(.+)(\s)(show version.txt)')
fileshowInv = re.compile(r'(.+)(\s)(show inventory.txt)')
fileshowIpIntBrief = re.compile(r'(.+)(\s)(show ip int b.txt)')

# Define dictionaries
devicesDictionary = {}

# Counters
interfaceLargest = 0
dnsLargest = 0
ntpLargest = 0
voicePortLargest = 0
#Functions
def grabby_README():
    print('   ____           _     _             _____           _')
    print('  / ___|_ __ __ _| |__ | |__  _   _  |_   _|__   ___ | |___ ')
    print(' | |  _|  __/ _  | |_ \| |_ \| | | |   | |/ _ \ / _ \| / __|')
    print(' | |_| | | | (_| | |_) | |_) | |_| |   | | (_) | (_) | \__ \ ')
    print('  \____|_|  \__,_|_.__/|_.__/ \__, |   |_|\___/ \___/|_|___/')
    print('                              |___/')
    print('Grabby Tools')
    print('Grabby Tools is a suite of programs created by Brian Wcisel and consists of the following tools.')
    print('')
    print('Grabby Config')
    print('This program connects to voice gateways via SSH or telnet and creates 5 text files based on 5 show commands,')
    print('show version, show ip int b, show run, show inventory and show interface. The filenames are automatically ')
    print('created based on the hostname of the device being discovered.')
    print('')
    print('Grabby Text')
    print('This program loops through all files in the directory it is placed in.  It first copies')
    print('all filenames into a list called filenames.  From there it walks each file name looking for')
    print('files that have show run in the filename and puts them into another list called hostnames.')
    print('The program then creates a dictionary with keys being hostnames and each value being another dictionary')
    print('with subsequent keys and values that correspond to all of the details captured by the regex-based text ')
    print('searches such as device hostname, default route, ntp, dns entries and interface details.  The output is then')
    print('written to a CSV with rows that represent each device by hostname and columns that represent the discovered ')
    print('data in a dynamic fashion. A formatted Spreadsheet is created from the finished CSV.')
    print('')
    print('Grabby DNS Checker')
    print('This program will check FQDNs and IP addresses in DNS to check for overlaps.')
    print('An input file (CSV) must be present in the directory the tool is run from.')
    print('The CSV has two columns with no headers.  First column is fqdns and the second column is IP addresses')
    print('Sometimes this tool will appear as if it is stalling on first run, this is because of DNS server response')
    print('Usually the program complete almost instantly after the first attempt.')
    time.sleep(60)

#Define Main functions

def grabby_dns_check(row):
    # Use Regex to parse row data
    # Feeds fow from input file into function for analysis
    if re.match(fqdnRE, row):
        # Store FQDN as variable stripping brackets, removing single quotes and forcing lowercase.
        fqdn = ((re.search(fqdnRE, row).group(1)).strip("[]'")).lower()
        # Store ip address as variable
        ipadd = ((re.search(fqdnRE, row).group(4)).strip("[]'")).lower()
        # Use built in Socket method to resolve fqdn
        try:
            resolved_ipaddr = socket.gethostbyname(fqdn)
        except socket.gaierror:
            DNS_failure_list.append("FORWARD_LOOKUP_FAILURE_RECORD_NOT_FOUND!,{}".format(fqdn))
            print("FORWARD_LOOKUP_FAILURE_RECORD_NOT_FOUND!,{}".format(fqdn))
        try:
            resolved_hostname = socket.gethostbyaddr(ipadd)
            # Split Tuple into three variables #tuple value 1 is the A Record first returned, tuple value 2 are any additional Aliases detected
            # tuple value 3 is IP Address
            recordA, aliasList, originalIP = resolved_hostname
            # Create a list of aliases to aliases to iterate over and document
            alias_list = []
            for i in aliasList:
                alias_list.append(i)
            originalIP = str(originalIP)
            originalIP = originalIP.strip("'[]")
            # Determine if there are aliases or duplicates returned with getipadrr, tuple 2 will be populated if so
            if is_empty(aliasList) is True:
                pass
            else:
                print("The_following_PTR_records_are_resolving_for_{}_when_there_should_only_be_one {}".format(ipadd, recordA))
                DNS_failure_list.append("The_following_PTR_records_are_resolving_for_{}_when_there_should_only_be_one {}".format(ipadd, recordA))
                for i in alias_list:
                    DNS_failure_list.append("-{}".format(i))
                    print("-{}".format(i))

        except socket.herror:
            print("REVERSE_LOOKUP_FAILED_no_PTR_Records_for {}".format(ipadd))
            DNS_failure_list.append("REVERSE_LOOKUP_FAILED_no_PTR_Records_for {}".format(ipadd))
        # Perform comparison
        try:
            if ipadd != resolved_ipaddr:
                print('\nDNS-A_Record_mismatch_for_{}_it_returns_{}_when_it_should_be {}'.format(fqdn, resolved_ipaddr, ipadd))
                DNS_failure_list.append('DNS-A_Record_mismatch_for_{}_it_returns_{}_when_it_should_be {}'.format(fqdn, resolved_ipaddr, ipadd))
        except:
            pass

def is_empty(tuple_value):
    # Determine if tuple value is empty.
    if tuple_value:
        # print('Structure is not empty.')
        return False
    else:
        # print('Structure is empty.')
        return True

# Write Discovered details to CSV then convert to XLSX for pretty formatting.
def spread_sheet_creation():
    try:
        with open('NetOutput.csv', 'w') as outputfile:
            # Create an object which operates like a regular writer but maps dictionaries onto Output rows
            writer = csv.DictWriter(outputfile, fieldnames=headers, lineterminator='\n')  # define writer csv using the fieldnames columns
            # Figure out how many headers for later XLSX Fill style operation
            column_count = len(headers)
            # Based on count, determine the column letter also used later for XLSX
            maxcolumns = (get_column_letter(column_count))
            # Write Headers
            writer.writeheader()
            # This is where the
            for i, j in devicesDictionary.items():
                #print(i, j)
                writer.writerow(j)
    except:
        print("Please Close the Netoutput.csv file and run the program again.")
        sys.exit()
    logging.info("Creating the Workbook took {} seconds".format(enddevicediscovery - startdevicediscovery))
    logging.info('############################################PROGRAM TERMINATED############################################')
    # open csv file
    csv_ = csv.reader(open('./NetOutput.csv'))
    reader = csv.reader(csv_, delimiter=",")
    # Create an Excel workbook object
    wb = Workbook()
    # Create an Excel worksheet object
    ws = wb.active
    # Give the worksheet a title
    ws.title = "NetOutput"
    # Go through each row in the csv in order to copy to XLSX with OPENPYXL library
    for ridx, row in enumerate(csv_):
        # Openpyxl starts row numbering at 1 so adjust the row index to match
        row_idx = ridx + 1
        # Go through each value in the csv row
        for cidx, val in enumerate(row):
            # Openpyxl starts column numbering at 1 so adjust the row index to match
            cell_idx = cidx + 1
            # Determine the excel cell name "A1", "B2", etc..
            # I have not seen this format used for a variable before.
            cell_name = '{}{}'.format(get_column_letter(cidx + 1), row_idx)
            # Create the cell object
            cell = ws[cell_name]
            # Set the value for the cell
            cell.value = escape_txt(val)
            # Set the cell format to text.  Don't ask my why @ means text but it does
            cell.number_format = '@'
    # Define fill object
    HeaderFill = PatternFill(fill_type='solid', fgColor='ff0000')
    # define font objects for later application
    bold_font = Font(bold=True)
    white_font = Font(color='FFFFFF', italic=False, bold=True)
    # Loop through the cells in the first column and apply the bold formatting
    for cell in ws['A:A']:
        cell.font = bold_font
    # Loop through the cells in the first row and apply the white text formatting
    for cell in ws["1:1"]:
        cell.font = white_font
    # Loop through the cells in the first row and apply the border and red fill formatting
    try:
        for row in ws['A1':maxcolumns]:
            for cell in row:
                cell.fill = HeaderFill
                cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                     right=Side(border_style='thin', color='FF000000'),
                                     bottom=Side(border_style='thin', color='FF000000'),
                                     left=Side(border_style='thin', color='FF000000'))
                # freeze pane.  Column and row.
                c = ws['B2']
                ws.freeze_panes = c
    except:
        print("Please Close the NetOutput file")
    # print('Writing Cell: {}, Value: {}, Format: {}'.format(cell_name, cell.value, cell.number_format))
    # Save the file
    try:
        wb.save(filename='NetOutput.xlsx')
    except:
        print("Please Close the NetOutput file")

def escape_txt(txt):
    # This function detects special characters and prefixes an apostrophe for to account for excel formating.
    special_chars = ['+', '=', '-', '/', '*', "'"]
    if txt:
        if txt[0] in special_chars:
            # print('Compensating (((((( {} ))))) by applying a prefix to accommodate excel formatting'.format(txt))
            logging.info('Compensating (((((( {} ))))) by applying a prefix to accommodate excel formatting'.format(txt))
            return "'" + txt
    return txt

def grabby_text_sh_ver1():
    try:
        logging.info('Opening file {} show Version.txt for analysis'.format(host))
        # Open Show Version based on Hostname
        with open(host + ' show version.txt') as showver:
            # print(host)
            # Iterate over each line.
            showversion = "Unknown"
            serial = "Unknown"
            lastreloadtype = "Unknown"
            lastreloadreason = "Unknown"
            lastreloadtime = "Unknown"
            configReg = "Unknown"
            modelType = "Unknown"
            nvRam = "Unknown"
            flashRam = "Unknown"
            for x in showver:
                # If line matches RE then bring it back to main function
                if re.match(showVerRE, x):
                    showversion = re.search(showVerRE, x).group(4)
                    # print(re.search(showVerRE, x).group(4))
                if re.match(serialNumRE, x):
                    serial = re.search(serialNumRE, x).group(1)
                    # print(re.search(serialNumRE, x).group(1))
                if re.match(lastReloadTypeRE, x):
                    lastreloadtype = re.search(lastReloadTypeRE, x).group(1)
                    # print(re.search(lastReloadTypeRE, x).group(1))
                if re.match(lastReloadReasonRE, x):
                    lastreloadreason = re.search(lastReloadReasonRE, x).group(1)
                if re.match(lastReloadTimeRE, x):
                    lastreloadtime = re.search(lastReloadTimeRE, x).group(1)
                    # print(re.search(lastReloadTimeRE, x).group(1))
                if re.match(configRegisterRE, x):
                    configReg = re.search(configRegisterRE, x).group(1)
                    # print(re.search(configRegisterRE, x).group(1))
                if re.match(modelTypeRE ,x):
                    modelType = re.search(modelTypeRE, x).group(2)
                    # print(re.search(modelTypeRE, x).group(2))
                if re.match(modelTypeRE2 ,x):
                    modelType = re.search(modelTypeRE2, x).group(2)
                    # print(re.search(modelTypeRE, x).group(2))
                if re.match(nvRamRE, x):
                    nvRam = re.search(nvRamRE, x).group(1)
                if re.match(nvRamRE2, x):
                    nvRam = re.search(nvRamRE2, x).group(1)
                if re.match(flashRamRE, x):
                    flashRam = re.search(flashRamRE, x).group(1)
                if re.match(flashRamRE2, x):
                    flashRam = re.search(flashRamRE2, x).group(1)
        return showversion, serial, lastreloadtype, lastreloadreason, lastreloadtime, configReg, modelType, nvRam, flashRam
    except:
        #print("fail")
        # print("There is no {} SHOW VERSION file".format(host))
        notProvided = 'Not Provided'
        showversion = notProvided
        serial = notProvided
        lastreloadtype = notProvided
        lastreloadreason = notProvided
        lastreloadtime = notProvided
        configReg = notProvided
        modelType = notProvided
        nvRam = notProvided
        flashRam = notProvided
        logging.info('There is no {} SHOW VERSION file'.format(host))
        return showversion, serial, lastreloadtype, lastreloadreason, lastreloadtime, configReg, modelType, nvRam, flashRam

def grabby_text_sh_int():
    try:
        logging.info('Opening file {} show Interface.txt for analysis'.format(host))
        with open(host + ' show interface.txt') as showint:
            for m in showint:
                if interfacex in m:
                    next_line = next(showint)
                    if re.search(showMacRE, next_line):
                        # Search for line for description RE
                        global macaddress
                        # Find the macaddress!
                        rawmac = re.search(showMacRE,  next_line).group(2)
                        # Get rid of the dots in the mac address format xxxx.xxxx.xxxx  MORE DOTS!  ...okay stop dots.
                        macaddress = rawmac.replace(".", "")
        return(macaddress)
    except:
        macaddress = "Not Provided"
        logging.info('There is no {} SHOW INTERFACE file'.format(host))
        # print("There is no {} SHOW INTERFACE file".format(host))
        return(macaddress)

def grabby_text_sh_run():
    # Define function dictionary
    deviceLocalDictionary = OrderedDict()
    # Open each show run based on hostname
    with open(host + ' show run.txt') as f:
        logging.info('Opening file {} show run.txt for analysis'.format(host))
        # Global Variables for one and all!  This is to allow these values to be referenced later in the program
        global interfaceLargest
        global dnsLargest
        global ntpLargest
        global voicePortLargest
        interfaceCounter = 0
        voicePortCounter = 0
        dnsCounter = 0
        ntpCounter = 0
        macCounter = 1
        CCMGROUPCounter = 0
        # Define default values for these values.  It helps to even have a value of nothing for these keys.
        deviceLocalDictionary["Default Route"] = ""
        deviceLocalDictionary["Domain Name"] = ""
        deviceLocalDictionary["SNMP"] = ""
        deviceLocalDictionary["SCCP CUCM SERVER 1"] = ""
        deviceLocalDictionary["SCCP CUCM SERVER 2"] = ""
        deviceLocalDictionary["SCCP CUCM SERVER 3"] = ""
        # See below at iproute search.  There must be a better way to do this.
        # search line by line in file
        for line in f:
            if re.match(hostnameRE, line):
                # Append dictionary with key hostname
                #  and value of matched regex group which is the actual hostname text
                deviceLocalDictionary["Hostname"] = re.search(hostnameRE, line).group(1)
                logging.debug('regex match for {} hostname '.format(re.search(hostnameRE, line).group(1)))
            if re.match(dnsRE, line):  # If DNS Server is found through matching regex regex
                dnsCounter += 1  # Increase counter to 1 if it is a zero or +1 if it is already (n)
                # Add key to deviceLocalDictionary dictionary called "DNS SERVER X" where X is the number of times this match
                #  was detected, DNS Server 1 X.X.X.X DNS Server 2 X.X.X.X etc.  The right side of this operation occurs
                # before the left.
                deviceLocalDictionary["DNS Server " + str(dnsCounter)] = re.search(dnsRE, line).group(1)
                logging.debug('regex match for {} {} DNS server '.format(host, re.search(dnsRE, line).group(1)))
                # Determine Largest counter.  Used for dynamic CSV header creation later on in the tool.
                if dnsCounter > dnsLargest:
                    dnsLargest = dnsCounter
                    logging.debug('DNS largest counter increased to {}'.format(dnsLargest))
            if re.match(ntpServerRE, line):
                # See above comment, identical Logic
                ntpCounter += 1
                deviceLocalDictionary["NTP Server " + str(ntpCounter)] = re.search(ntpServerRE, line).group(1)
                logging.debug('regex match for {} {} NTP server '.format(host, re.search(ntpServerRE, line).group(1)))
                if ntpCounter > ntpLargest:
                    ntpLargest = ntpCounter
                    logging.debug('NTP largest counter increased to {}'.format(ntpCounter))
            # There has to be a better way to represent absent values than a boolean check... I'monly doing this so I can place a blank
            if re.match(ipRouteRE, line):
                deviceLocalDictionary["Default Route"] = re.search(ipRouteRE, line).group(1)
                logging.debug('regex match for {} Default Route - "{}" '.format(host, re.search(ipRouteRE, line).group(1)))
            if re.match(snmpRE, line):
                deviceLocalDictionary["SNMP"] = re.search(snmpRE, line).group(1)
                logging.debug('regex match for {} SNMP string - "{}"'.format(host, re.search(snmpRE, line).group(1)))
            if re.match(ipDomainRE, line):
                deviceLocalDictionary["Domain Name"] = re.search(ipDomainRE, line).group(1)
                logging.debug('regex match for {} domain name - "{}" '.format(host, re.search(ipDomainRE, line).group(1)))
            # These lists are used to to store interface and voice port sub-commands for subsequent analysis
            intDetailList = []
            voiceportDetailList = []
            if re.match(sccpRegex, line):
                CCMGROUPCounter += 1
                # debugging
                # test_var = re.search(sccpRegex, line).group(1)
                # print(test_var)
                deviceLocalDictionary["SCCP CUCM SERVER " + str(CCMGROUPCounter)] = re.search(sccpRegex, line).group(1)

                logging.debug('regex match for {} {} SCCP CUCM server '.format(host, re.search(sccpRegex, line).group(1)))
            # Begin Interface Parsing
            if re.match(interfacetypeRE, line):  # If Regex hit true then
                # Define global interfacex as the match criteria for finding mac address in show interface output
                global interfacex
                interfacex = re.search(interfacetypeRE, line).group(1)
                logging.debug('regex match for {} interface - "{}"'.format(host, re.search(interfacetypeRE, line).group(1)))
                # call function to match interface type with its mac address.  This requires the show int file
                macccheck = grabby_text_sh_int() # TODO rename this to something that makes sense!
                deviceLocalDictionary["MAC Address {}".format(macCounter)] = macccheck
                logging.debug('regex MAC ADDRESS match for {} {} - "{}" '.format(host, re.search(interfacetypeRE, line).group(1), macccheck))
                interfaceCounter += 1
                macCounter += 1
                # Search line for Hit matching Interface-type regex pattern defined above
                deviceLocalDictionary["Interface {}".format(interfaceCounter)] = re.search(interfacetypeRE, line).group(1)
                # Set highest highest interface counter by comparison.  This value is used to determine how many headers to write later on
                if interfaceCounter > interfaceLargest:
                    interfaceLargest = interfaceCounter
                    logging.debug('Largest Interface Counter increased to {}'.format(interfaceLargest))
                intDetailList.append(line)
                # After matching the interface type, take all of the subcommands and put them into a list, stopping at !, like a search terminator
                while line != str("!\n"):
                    line = next(f)
                    intDetailList.append(line)
                    # Strip new lines and leading or trailing spaces, item by item in the list.
                    intDetailList[:] = [line.rstrip('\n') for line in intDetailList]
                    intDetailList[:] = [line.lstrip() for line in intDetailList]
                # Define my null values for use in the finished NETOUTPUT.CSV  This is like zeroing out the values
                deviceLocalDictionary["Description {}".format(interfaceCounter)] = ""
                deviceLocalDictionary["Ip Address {}".format(interfaceCounter)] = ""
                deviceLocalDictionary["Subnet Mask {}".format(interfaceCounter)] = ""
                deviceLocalDictionary["Channel Group {}".format(interfaceCounter)] = ""
                # Crawl over the interface list looking for specific details.  Overwrite above keys when/if found
                for item in intDetailList:
                    if re.search(intDescRE, item):
                        deviceLocalDictionary["Description {}".format(interfaceCounter)] = re.search(intDescRE, item).group(1)
                    if re.search(ipAddRE, item):
                        deviceLocalDictionary["Ip Address {}".format(interfaceCounter)] = re.search(ipAddRE, item).group(1)
                        deviceLocalDictionary["Subnet Mask {}".format(interfaceCounter)] = re.search(ipAddRE, item).group(2)
                    if re.search(ipChanGrpRE, item):
                        deviceLocalDictionary["Channel Group {}".format(interfaceCounter)] = "Channel Group {}".format(re.search(ipChanGrpRE, item).group(2))
            if re.match(voicePortRE, line):
                logging.debug('regex match for - "{}"'.format(re.search(voicePortRE, line).group(1)))
                voicePortCounter += 1
                if voicePortCounter > voicePortLargest:
                    voicePortLargest = voicePortCounter
                    logging.debug('Largest Voice Port Counter increased to {}'.format(voicePortLargest))
                deviceLocalDictionary["Voice Port {}".format(voicePortCounter)] = re.search(voicePortRE, line).group(1)
                while line != str("!\n"):
                    line = next(f)
                    voiceportDetailList.append(line)
                    # Strip new lines and leading or trailing spaces, item by item in the list.
                    # voiceportDetailList[:] = [line.rstrip('\n') for line in voiceportDetailList]
                    # voiceportDetailList[:] = [line.lstrip() for line in voiceportDetailList]
                deviceLocalDictionary["Voice Port Description {}".format(voicePortCounter)] = ""
                for item in voiceportDetailList:
                    if re.search(intDescRE, item):
                        deviceLocalDictionary["Voice Port Description {}".format(voicePortCounter)] = re.search(intDescRE,item).group(1)

        showversion = grabby_text_sh_ver1()
        deviceLocalDictionary["Software Version"] = showversion[0]
        deviceLocalDictionary["Serial Number"] = showversion[1]
        deviceLocalDictionary["Last Reload Type"] = showversion[2]
        deviceLocalDictionary["Last Reload Reason"] = "{} at {}".format(showversion[3], showversion[4])
        deviceLocalDictionary["Config Register"] = showversion[5]
        deviceLocalDictionary["Model Type"] = showversion[6]
        deviceLocalDictionary["NVRAM"] = showversion[7]
        deviceLocalDictionary["Flash"] = showversion[8]


    return deviceLocalDictionary

def grabby_config_devicediscovery_all_details(netdata):
    # FUNCTION DEFINITION: This function writes show commands to output files
    try:
        logging.info('Beginning Grabby Config Device Discovery')
        # Import Timing
        from timeit import default_timer as timer
        # Define performance timer
        startdevicediscovery = timer()
        print("\n")
        logging.info("Attempting to connect to {} in order to build Show Command Outputs".format(netdata[ipaddress]))
        print("Attempting to connect to {} in order to build Show Command Outputs".format(netdata[ipaddress]))
        # SSH Connection Tools from Netmiko library
        from netmiko import ConnectHandler

        # Pass data to Connection Handler
        establish_session = ConnectHandler(device_type=netdata[devicetype], ip=netdata[ipaddress], username=netdata[username], password=netdata[password], secret=netdata[secret])

        # Pass commands to CLI through existing SSH session
        show_run = establish_session.send_command('show run')
        show_ip_int_b = establish_session.send_command('show ip int b')
        show_inventory = establish_session.send_command('show inventory')
        show_version = establish_session.send_command('show version')
        show_interface = establish_session.send_command('show interface')
        # Search running config for hostname and use that as output file name.
        hostname = re.search(r'hostname\s(.+)', show_run)
        logging.info("Files for  {} successfully created ".format(hostname.group(1)))

        # Create output show run text file
        show_run_output = open(hostname.group(1) + " show run.txt", "w")
        show_ip_int_b_output = open(hostname.group(1) + " show ip int b.txt", "w")
        show_inventory_output = open(hostname.group(1) + " show inventory.txt", "w")
        show_version_output = open(hostname.group(1) + " show version.txt", "w")
        show_interface_output = open(hostname.group(1) + " show interface.txt", "w")

        # Write show run text to output file
        show_run_output.write(show_run)
        show_ip_int_b_output.write(show_ip_int_b)
        show_inventory_output.write(show_inventory)
        show_version_output.write(show_version)
        show_interface_output.write(show_interface)

        # Close file show run file
        show_run_output.close()
        show_ip_int_b_output.close()
        show_inventory_output.close()
        show_version_output.close()
        show_interface_output.close()

        print("Connecting to {} via SSH was successful".format(netdata[ipaddress]))
        logging.info("Connecting to {} via SSH was successful".format(netdata[ipaddress]))
        # End Performance Timer
        enddevicediscovery = timer()
        logging.info("Discovering details from this device took {} seconds".format(round(enddevicediscovery - startdevicediscovery)))
        print("Discovering details from this device took {} seconds".format(round(enddevicediscovery - startdevicediscovery)))

    except:
        try:
            print("Connection to host at Ip Address {} via SSH failed, now attempting Telnet session".format(netdata[ipaddress]))
            # Pass data to Connection Handler
            establish_session = ConnectHandler(device_type='cisco_ios_telnet', ip=netdata[ipaddress], password=netdata[password], secret=netdata[secret])
            # The below command is required for telnet sessions.  Netmiko sends the enable command to enter global config.
            establish_session.enable()
            # Pass commands to CLI through existing SSH session
            show_run = establish_session.send_command('show run')
            show_ip_int_b = establish_session.send_command('show ip int b')
            show_inventory = establish_session.send_command('show inventory')
            show_version = establish_session.send_command('show version')
            show_interface = establish_session.send_command('show interface')
            # Search running config for hostname and use that as output file name.
            hostname = re.search(r'hostname\s(.+)', show_run)

            # Create output show run text file
            show_run_output = open(hostname.group(1) + " show run.txt", "w")
            show_ip_int_b_output = open(hostname.group(1) + " show ip int b.txt", "w")
            show_inventory_output = open(hostname.group(1) + " show inventory.txt", "w")
            show_version_output = open(hostname.group(1) + " show version.txt", "w")
            show_interface_output = open(hostname.group(1) + " show interface.txt", "w")

            # Write show run text to output file
            show_run_output.write(show_run)
            show_ip_int_b_output.write(show_ip_int_b)
            show_inventory_output.write(show_inventory)
            show_version_output.write(show_version)
            show_interface_output.write(show_interface)

            # Close file show run file
            show_run_output.close()
            show_ip_int_b_output.close()
            show_inventory_output.close()
            show_version_output.close()
            show_interface_output.close()


            print("Connecting to {} via Telnet was successful".format(netdata[ipaddress]))
            logging.info("Connecting to {} via Telnet was successful".format(netdata[ipaddress]))
            # End Performance Timer
            enddevicediscovery = timer()
            logging.info("Discovering details from this device took {} seconds".format(round(enddevicediscovery - startdevicediscovery)))
            print("Discovering details from this device took {} seconds".format(round(enddevicediscovery - startdevicediscovery)))
        except:
            print("Connection to host at Ip Address {} via Telnet failed.".format(netdata[ipaddress]))
            print("Try updating the input file IOS type to read 'cisco_ios_telnet' and rerun.  Errors may be misleading for telnet operation")
            # TODO Add failed things here.
            # print('Except')
            # print(netdata[ipaddress])
            failed_connections_list.append(netdata[ipaddress])
            # print(failed_connections_list)


def grabby_config_devicediscovery_show_run_only(netdata):
    # FUNCTION DEFINITION: This function writes show commands to output files
    try:
        logging.info('Beginning Grabby Config Device Discovery')
        # Import Timing
        from timeit import default_timer as timer
        # Define performance timer
        startdevicediscovery = timer()
        print("\n")
        logging.info("Attempting to connect to {} in order to build Show Run Outputs".format(netdata[ipaddress]))
        print("Attempting to connect to {} in order to build Show Run Outputs".format(netdata[ipaddress]))
        # SSH Connection Tools from Netmiko library
        from netmiko import ConnectHandler

        # Pass data to Connection Handler
        establish_session = ConnectHandler(device_type=netdata[devicetype], ip=netdata[ipaddress], username=netdata[username], password=netdata[password], secret=netdata[secret])

        # Pass commands to CLI through existing SSH session
        show_run = establish_session.send_command('show run')

        hostname = re.search(r'hostname\s(.+)', show_run)
        logging.info("Files for  {} successfully created ".format(hostname.group(1)))

        # Create output show run text file
        show_run_output = open(hostname.group(1) + " show run.txt", "w")


        # Write show run text to output file
        show_run_output.write(show_run)


        # Close file show run file
        show_run_output.close()


        print("Connecting to {} via SSH was successful".format(netdata[ipaddress]))
        logging.info("Connecting to {} via SSH was successful".format(netdata[ipaddress]))
        # End Performance Timer
        enddevicediscovery = timer()
        logging.info("Discovering details from this device took {} seconds".format(round(enddevicediscovery - startdevicediscovery)))
        print("Discovering details from this device took {} seconds".format(round(enddevicediscovery - startdevicediscovery)))

    except:
        try:
            print("Connection to host at Ip Address {} via SSH failed, now attempting Telnet session".format(netdata[ipaddress]))
            # Pass data to Connection Handler
            establish_session = ConnectHandler(device_type='cisco_ios_telnet', ip=netdata[ipaddress], password=netdata[password], secret=netdata[secret])
            # The below command is required for telnet sessions.  Netmiko sends the enable command to enter global config.
            establish_session.enable()
            # Pass commands to CLI through existing SSH session
            show_run = establish_session.send_command('show run')


            # Create output show run text file
            show_run_output = open(hostname.group(1) + " show run.txt", "w")

            # Write show run text to output file
            show_run_output.write(show_run)

            # Close file show run file
            show_run_output.close()


            print("Connecting to {} via Telnet was successful".format(netdata[ipaddress]))
            logging.info("Connecting to {} via Telnet was successful".format(netdata[ipaddress]))
            # End Performance Timer
            enddevicediscovery = timer()
            logging.info("Discovering details from this device took {} seconds".format(round(enddevicediscovery - startdevicediscovery)))
            print("Discovering details from this device took {} seconds".format(round(enddevicediscovery - startdevicediscovery)))
        except:
            print("Connection to host at Ip Address {} via Telnet failed.".format(netdata[ipaddress]))
            print("Try updating the input file IOS type to read 'cisco_ios_telnet' and rerun.  Errors may be misleading for telnet operation")
            # TODO Add failed things here.
            # print(netdata[ipaddress])
            failed_connections_list.append(netdata[ipaddress])

logging.info('############################################DISCOVERING FILES IN DIRECTORY############################################')


 #################################PROGRAM START after functions are loaded###############################################################
 # loop through hostnames list and initialize devicesDictionary

# Define list if directories in current working directory for filename walk.  This determines which files are present
hostnames = []
filenames = [f for f in listdir()]
# iterate over each file and put discovered hostnames captured from filewalk into the list titled 'hostnames'.
for file in filenames:
    logging.debug('Filename {} detected'.format(file))
    if re.match(showRunRE, file):
        logging.info('Valid Config Filename {} detected'.format(file))
        hostnames.append(re.search(showRunRE, file).group(1))
    if re.match(showIntRE1, file):
        logging.info('Valid Config Filename {} detected'.format(file))
    if re.match(fileshowVerRE2, file):
        logging.info('Valid Config Filename {} detected'.format(file))
    if re.match(fileshowInv, file):
        logging.info('Valid Config Filename {} detected'.format(file))
    if re.match(fileshowIpIntBrief, file):
        logging.info('Valid Config Filename {} detected'.format(file))



########################################### Ask User to Enter Selection ###########################################
# 1st Prompt Ask user which tools to run
print("Input '1' to run GRABBY CONFIG to discovers ALL CONFIG Files from devices defined in Netinput.csv")
print("Input '2' to run GRABBY CONFIG to discover only SHOW RUN files from devices defined in Netinput.csv")
print("Input '3' to run GRABBY TEXT which creates a Workbook CSV from recently discovered files")
print("Input '4' to run GRABBY DNS Checker which checks a CSV file of A and PTR records for overlaps or duplicates")
print("Input '5' for README")
selection = input("What is your selection?: ")

if selection == str(1):
    print("\n")
    print("   ______ ____   ___     ____   ____ __  __   ______ ____   _   __ ______ ____ ______")
    print("  / ____// __ \ /   |   / __ ) / __ )\ \/ /  / ____// __ \ / | / // ____//  _// ____/")
    print(" / / __ / /_/ // /| |  / __  |/ __  | \  /  / /    / / / //  |/ // /_    / / / / __  ")
    print("/ /_/ // _, _// ___ | / /_/ // /_/ /  / /  / /___ / /_/ // /|  // __/  _/ / / /_/ /")
    print("\____//_/ |_|/_/  |_|/_____//_____/  /_/   \____/ \____//_/ |_//_/    /___/ \____/")
    print("\n")



    logging.info('############################################GRABBY CONFIG STARTED############################################')
    logging.info('Option 1 Selected')

    try:
        inputfile = open('NetInput.csv', 'rt')
    except:
        print("\n")
        print("ERROR - Netinput.csv was not detected please add the file and run GrabbyTools again.")
        print("Program will terminate in ten seconds")
        logging.info('Netinput.csv not detected.  Program terminating.')
        time.sleep(10)
        sys.exit()

    # Creating new Directory
    newDirectory = time.strftime("%d_%B_%Y_%I_%M_%S_Grabby_Data_Output")
    os.makedirs(newDirectory)
    os.chdir(newDirectory)
    # Open file for failed connection attempts HEY!  DO SOMETHING WITH THIS!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    failedConnections = open('Failed_Device_Connections.txt', 'w')
    # TODO add csv output of failed connection results.  Make this look like the input file.
    rownum = 0
    # creates the reader object from CSV file
    reader = csv.reader(inputfile)
    # Main Program Performance Timer
    from timeit import default_timer as timer
    startglobaltimer = timer()
    start = timer()
    # Defines Headers to match CSV headers
    list_threads = []
    for row in reader:
        if rownum == 0:
            header = row
            devicetype = header.index('Device Type')
            ipaddress = header.index('IPAddress')
            username = header.index('Username')
            password = header.index('Password')
            secret = header.index('Secret')
            discovery = header.index('Discovery')
        else:
            colnum = 0
            for col in row:
                if (colnum == header.index('Discovery')) and (col.lower() == 'y'):
                    netdata = row
                    t1 = threading.Thread(target=grabby_config_devicediscovery_all_details, args=(netdata,))
                    list_threads.append(t1)

                    try:
                        # time.sleep(.1)
                        # t1.stack_size([1])
                        # TODO Figure out how to make the above command work to throttle threads
                        t1.start()
                    except:
                        print("Cannot Run Threading Operation.  Check connectivity and input file")
                        pass
                colnum += 1  # Loop Counter
        rownum += 1  # Loop Counter
    end = timer()
    #Close failed connections txt
    for t in list_threads:
        try:
            t1.join()
        except:
            print("Failed interation")
d
    # print(failed_connections_list)
    for i in failed_connections_list:
        failedConnections.write(i + "\n")
    failedConnections.close()
    print("\n")
    #print("The overall operation took {} seconds".format(round(end - start)))
    # TODO Figure out why threading breaks the above timer for over all performance
    # logging.info('############################################PROGRAM TERMINATED############################################')
    # TODO Figure out why threading breaks logging, probably related to the above threading issue as well.

elif selection == str(2):
    #
    print("\n")
    print("   ______ ____   ___     ____   ____ __  __   ______ ____   _   __ ______ ____ ______")
    print("  / ____// __ \ /   |   / __ ) / __ )\ \/ /  / ____// __ \ / | / // ____//  _// ____/")
    print(" / / __ / /_/ // /| |  / __  |/ __  | \  /  / /    / / / //  |/ // /_    / / / / __  ")
    print("/ /_/ // _, _// ___ | / /_/ // /_/ /  / /  / /___ / /_/ // /|  // __/  _/ / / /_/ /")
    print("\____//_/ |_|/_/  |_|/_____//_____/  /_/   \____/ \____//_/ |_//_/    /___/ \____/")
    print("\n")



    logging.info('############################################GRABBY CONFIG STARTED############################################')
    logging.info('Option 2 Selected')

    try:
        inputfile = open('NetInput.csv', 'rt')
    except:
        print("\n")
        print("ERROR - Netinput.csv was not detected please add the file and run GrabbyTools again.")
        print("Program will terminate in ten seconds")
        logging.info('Netinput.csv not detected.  Program terminating.')
        time.sleep(10)
        sys.exit()

    # Creating new Directory
    newDirectory = time.strftime("%d_%B_%Y_%I_%M_%S_Grabby_Data_Output")
    os.makedirs(newDirectory)
    os.chdir(newDirectory)
    # Open file for failed connection attempts HEY!  DO SOMETHING WITH THIS!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    failedConnections = open('Failed_Device_Connections.txt', 'w')
    # TODO add csv output of failed connection results.  Make this look like the input file.
    rownum = 0
    # creates the reader object from CSV file
    reader = csv.reader(inputfile)
    # Main Program Performance Timer
    from timeit import default_timer as timer
    startglobaltimer = timer()
    start = timer()
    # Defines Headers to match CSV headers
    for row in reader:
        if rownum == 0:
            header = row
            devicetype = header.index('Device Type')
            ipaddress = header.index('IPAddress')
            username = header.index('Username')
            password = header.index('Password')
            secret = header.index('Secret')
            discovery = header.index('Discovery')
        else:
            colnum = 0
            for col in row:
                if (colnum == header.index('Discovery')) and (col.lower() == 'y'):
                    netdata = row
                    t1 = threading.Thread(target=grabby_config_devicediscovery_show_run_only, args=(netdata,))
                    try:
                        # time.sleep(.1)
                        # t1.stack_size([1])
                        # TODO Figure out how to make the above command work to throttle threads
                        t1.start()
                    except:
                        print("Cannot Run Threading Operation.  Check connectivity and input file")
                        pass
                colnum += 1  # Loop Counter
        rownum += 1  # Loop Counter
    end = timer()
    #Close failed connections txt
    for i in failed_connections_list:
        # print("hit 2")
        failedConnections.write(i)
    failedConnections.close()
    print("\n")
    #print("The overall operation took {} seconds".format(round(end - start)))
    # TODO Figure out why threading breaks the above timer for over all performance
    logging.info('############################################PROGRAM TERMINATED############################################')

elif selection == str(3):
    print("\n")
    print("   __________  ___    ____  ______  __   _____________  ________")
    print("  / ____/ __ \/   |  / __ )/ __ ) \/ /  /_  __/ ____/ |/ /_  __/")
    print(" / / __/ /_/ / /| | / __  / __  |\  /    / / / __/  |   / / /   ")
    print("/ /_/ / _, _/ ___ |/ /_/ / /_/ / / /    / / / /___ /   | / /    ")
    print("\____/_/ |_/_/  |_/_____/_____/ /_/    /_/ /_____//_/|_|/_/  ")
    print("\n")
    logging.info('############################################GRABBY TEXT STARTED############################################')
    logging.info('Option 2 Selected')
    # Main Program Performance Timer
    # create dictionary for easy selection display
    subdirdict = OrderedDict()
    # create list for later selection
    subdirlist = []
    # counter
    subdircounter = 0

    # Detect if there are files in the current directory
    if glob('*show run.txt'):
        from timeit import default_timer as timer
        # Define performance timer
        startdevicediscovery = timer()
        for host in hostnames:
            devicesDictionary[host] = grabby_text_sh_run()
        enddevicediscovery = timer()
        #Arbitrarily apply Hostname Header as the first header in the list of headers, which are created in the next section
        headers = ["Hostname"]
        # Create Headers
        # Here we check to see the highest number of DNS entries discovered so that we can write a maximum of x headers.
        # These are written out as DNS Server 1, DNS Server 2, ... DNS Server X in the final output
        for number in range(0, int(dnsLargest)):
            headers.append("DNS Server " + str(number + 1))
        #Same logic as DNS, we're determining the amount of ntp server headers to write
        for number in range(0, int(ntpLargest)):
            headers.append("NTP Server " + str(number + 1))
        # Arbitrarily create headers for our output like above with hostname
        headers.append('Domain Name')
        headers.append('SNMP')
        headers.append('Default Route')
        headers.append('Serial Number')
        headers.append('Software Version')
        headers.append('SCCP CUCM SERVER 1')
        headers.append('SCCP CUCM SERVER 2')
        headers.append('SCCP CUCM SERVER 3')
        headers.append('Last Reload Type')
        headers.append('Last Reload Reason')
        headers.append('Config Register')
        headers.append('Model Type')
        headers.append('NVRAM')
        headers.append('Flash')
        # Header creation, based on hightest number of interface hits on any gateway discovered.
        for number in range(0, int(interfaceLargest)):
            headers.append("Interface {}".format((number + 1)))
            headers.append("Description " + str(number + 1))
            headers.append("Ip Address " + str(number + 1))
            headers.append("Subnet Mask " + str(number + 1))
            headers.append("MAC Address " + str(number + 1))
            headers.append("Channel Group " + str(number + 1))
        # Header creation, based on highest number of voice port hits on any gateway discovered.
        for number in range(0, int(voicePortLargest)):
            headers.append("Voice Port " + str(number + 1))
            headers.append("Voice Port Description " + str(number + 1))
        headercount = (len(headers))
        spread_sheet_creation()

        print("Operation Complete!")
        print("")
        print("Creating the Workbook took {} seconds".format(round(enddevicediscovery - startdevicediscovery, 4)))
        print("This operation took {} seconds".format(round(enddevicediscovery - startdevicediscovery, 9)))
    #     ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    else:

        # Search current directory and make a list of folders
        for dir in glob('*/'):
            #append the list used for user input later
            subdirlist.append(dir)
            subdirlist[:] = [line.rstrip('\n') for line in subdirlist]
            # Counter to create unique entries
            subdircounter = subdircounter + 1
            subdirdict[(subdircounter)] = dir
        print("These are the folders discovered in this directory")
        for k, v in subdirdict.items():
            print(k, v)
        print("")
        # Ask user which directory they want to run the
        selection = input("Which directory? (Enter Numeral): ")
        print("")
        selection = int(selection)
        #validation that the selection is within the number or detected directories
        if int(selection) <= int(subdircounter):
            # adjust number down one
            chosen_directory = subdirlist[int(selection) - 1]
            # print(os.getcwd() + "\\" + chosen_directory)
            # new directory path
            new_directory_path = os.getcwd() + "\\" + chosen_directory
            # Change the current working directory
            os.chdir(new_directory_path)
            logging.debug('Current Working Directory changed to - {}'.format(new_directory_path))
        else:
            print("Bad selection")

        filenames = [f for f in listdir()]
        # iterate over each file and put discovered hostnames captured from filewalk into the list titled 'hostnames'.
        for file in filenames:
            logging.debug('Filename {} detected'.format(file))
            if re.match(showRunRE, file):
                logging.info('Valid Config Filename {} detected'.format(file))
                hostnames.append(re.search(showRunRE, file).group(1))
            if re.match(showIntRE1, file):
                logging.info('Valid Config Filename {} detected'.format(file))
            if re.match(fileshowVerRE2, file):
                logging.info('Valid Config Filename {} detected'.format(file))
            if re.match(fileshowInv, file):
                logging.info('Valid Config Filename {} detected'.format(file))
            if re.match(fileshowIpIntBrief, file):
                logging.info('Valid Config Filename {} detected'.format(file))
        from timeit import default_timer as timer

        # Define performance timer
        startdevicediscovery = timer()
        for host in hostnames:
            devicesDictionary[host] = grabby_text_sh_run()
        enddevicediscovery = timer()
        headers = ["Hostname"]
        # Create Headers
        # Determine number of headers to write to CSV

        for number in range(0, int(dnsLargest)):
            headers.append("DNS Server " + str(number + 1))
        for number in range(0, int(ntpLargest)):
            headers.append("NTP Server " + str(number + 1))
        headers.append('Domain Name')
        headers.append('SNMP')
        headers.append('Default Route')
        headers.append('Serial Number')
        headers.append('Software Version')
        headers.append('SCCP CUCM SERVER 1')
        headers.append('SCCP CUCM SERVER 2')
        headers.append('SCCP CUCM SERVER 3')
        headers.append('Last Reload Type')
        headers.append('Last Reload Reason')
        headers.append('Config Register')
        headers.append('Model Type')
        headers.append('NVRAM')
        headers.append('Flash')
        for number in range(0, int(interfaceLargest)):
            headers.append("Interface {}".format((number + 1)))
            headers.append("Description " + str(number + 1))
            headers.append("Ip Address " + str(number + 1))
            headers.append("Subnet Mask " + str(number + 1))
            headers.append("MAC Address " + str(number + 1))
            headers.append("Channel Group " + str(number + 1))
        for number in range(0, int(voicePortLargest)):
            headers.append("Voice Port " + str(number + 1))
            headers.append("Voice Port Description " + str(number + 1))
        # get header count for formatting later on
        headercount = (len(headers))
        spread_sheet_creation()

        print("Operation Complete!")
        print("")
        print("Creating the Workbook took {} seconds".format(round(enddevicediscovery - startdevicediscovery, 4)))
        print("This operation took {} seconds".format(round(enddevicediscovery - startdevicediscovery, 9)))

elif selection == str(4):

    print('   __________  ___    ____  ______  ______  _   _______ ________  ________________ __ __________ ')
    print('  / ____/ __ \/   |  / __ )/ __ ) \/ / __ \/ | / / ___// ____/ / / / ____/ ____/ //_// ____/ __ \ ')
    print(' / / __/ /_/ / /| | / __  / __  |\  / / / /  |/ /\__ \/ /   / /_/ / __/ / /   / ,<  / __/ / /_/ /')
    print('/ /_/ / _, _/ ___ |/ /_/ / /_/ / / / /_/ / /|  /___/ / /___/ __  / /___/ /___/ /| |/ /___/ _, _/')
    print('\____/_/ |_/_/  |_/_____/_____/ /_/_____/_/ |_//____/\____/_/ /_/_____/\____/_/ |_/_____/_/ |_|')
    print("")
    print("")


    ###############Important Note!########################
    # Sometimes DNS lookups can take a while to get started.  This tool may take a while when it is run the first time.
    # Subsequent attempts to run the program should work much quicker than the first attempt.
    # Remember to Flush DNS occasionally and if you need to check against a certain DNS server, statically
    # assign the server to your network connection.

    # Define Function to check if tuple value is empty or not.


    # Regular expressions for FQDN search
    fqdnRE = re.compile(r'(.+)(,)(\s)(.+)')

    # Open the file. Requires two columns, first column has FQDNs and second has IP addresses
    inputfile = open('DNSInput.csv', 'rt')

    # Create CSV Reader Object
    reader = csv.reader(inputfile)
    # Define failure list to store errors.  Eventually this list is written to a Failed jobs output file
    DNS_failure_list = []
    # Define a list to store each thread
    threads = []
    for row in reader:
        # Make row strings
        row = str(row)
        # Reset values
        resolved_ipaddr = "Unknown"
        resolved_hostname = "Unknown"
        # define local sets
        resolved_ipaddresses_set = set()
        resolved_hostnames_set = set()
        # Define the damn thread
        t1 = threading.Thread(target=grabby_dns_check, args=(row,))
        # Start the threading process
        t1.start()
        # Add individual threads to thread list.
        threads.append(t1)
    for proc in threads:
        # Join Threads.  This will run all threads at once and will move on once completed.  Better than Active Count
        t1.join()
    # print("\nAll records, or, all remaining records are valid")
    with open("GrabbyDNSCheck Failures.csv", 'w', newline='') as file:
        wr = csv.writer(file, dialect='excel')
        for i in DNS_failure_list:
            wr.writerow([i])

elif selection == str(5):
    grabby_README()
else:
    logging.error('Invalid Selection')
    print("\n")
    print("Invalid Selection")
    logging.info('############################################PROGRAM TERMINATED############################################')